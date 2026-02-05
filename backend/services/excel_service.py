"""
Service d'export Excel avec openpyxl
Crée des fichiers Excel professionnels avec formatage, graphiques, formules
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from typing import List, Dict, Any
import io


class ExcelService:
    """Service d'export Excel professionnel"""
    
    def __init__(self, title: str = "Export SIU"):
        self.title = title
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Retirer la feuille par défaut
        
        # Styles prédéfinis
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='673AB7', end_color='673AB7', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_sheet(self, title: str, activate: bool = True):
        """Crée une nouvelle feuille"""
        sheet = self.wb.create_sheet(title=title)
        if activate:
            self.wb.active = sheet
        return sheet
    
    def export_parcels(
        self,
        parcels: List[Dict[str, Any]],
        sheet_name: str = "Parcelles"
    ):
        """
        Exporte une liste de parcelles avec formatage
        
        Args:
            parcels: Liste des parcelles
            sheet_name: Nom de la feuille
        """
        sheet = self.create_sheet(sheet_name)
        
        # En-têtes
        headers = [
            'ID', 'Référence', 'Adresse', 'Superficie (m²)', 
            'Zone', 'Statut', 'Propriétaire', 'Date création'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # Données
        for row_num, parcel in enumerate(parcels, 2):
            sheet.cell(row=row_num, column=1).value = parcel.get('id')
            sheet.cell(row=row_num, column=2).value = parcel.get('reference_cadastrale')
            sheet.cell(row=row_num, column=3).value = parcel.get('address')
            sheet.cell(row=row_num, column=4).value = parcel.get('area')
            sheet.cell(row=row_num, column=5).value = parcel.get('zone')
            sheet.cell(row=row_num, column=6).value = parcel.get('status')
            sheet.cell(row=row_num, column=7).value = parcel.get('owner_name', 'N/A')
            
            # Date formatée
            created_at = parcel.get('created_at')
            if created_at:
                if isinstance(created_at, str):
                    created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                sheet.cell(row=row_num, column=8).value = created_at
                sheet.cell(row=row_num, column=8).number_format = 'DD/MM/YYYY'
            
            # Bordures
            for col in range(1, 9):
                sheet.cell(row=row_num, column=col).border = self.border
        
        # Auto-ajuster les colonnes
        for col in range(1, 9):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Ajouter des filtres
        sheet.auto_filter.ref = f"A1:H{len(parcels) + 1}"
        
        # Ligne de total pour superficie
        total_row = len(parcels) + 2
        sheet.cell(row=total_row, column=3).value = "TOTAL:"
        sheet.cell(row=total_row, column=3).font = Font(bold=True)
        
        sheet.cell(row=total_row, column=4).value = f"=SUM(D2:D{len(parcels) + 1})"
        sheet.cell(row=total_row, column=4).font = Font(bold=True)
    
    def export_documents(
        self,
        documents: List[Dict[str, Any]],
        sheet_name: str = "Documents"
    ):
        """Exporte une liste de documents"""
        sheet = self.create_sheet(sheet_name)
        
        headers = [
            'ID', 'Titre', 'Type', 'Taille', 'Statut', 
            'Uploadé par', 'Date upload', 'Validé'
        ]
        
        # En-têtes avec style
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Données
        for row_num, doc in enumerate(documents, 2):
            sheet.cell(row=row_num, column=1).value = doc.get('id')
            sheet.cell(row=row_num, column=2).value = doc.get('title')
            sheet.cell(row=row_num, column=3).value = doc.get('document_type')
            
            # Taille formatée
            size = doc.get('file_size', 0)
            size_mb = size / (1024 * 1024) if size else 0
            sheet.cell(row=row_num, column=4).value = f"{size_mb:.2f} MB"
            
            sheet.cell(row=row_num, column=5).value = doc.get('status')
            sheet.cell(row=row_num, column=6).value = doc.get('username', 'N/A')
            
            # Date
            uploaded_at = doc.get('uploaded_at')
            if uploaded_at:
                if isinstance(uploaded_at, str):
                    uploaded_at = datetime.fromisoformat(uploaded_at.replace('Z', '+00:00'))
                sheet.cell(row=row_num, column=7).value = uploaded_at
                sheet.cell(row=row_num, column=7).number_format = 'DD/MM/YYYY HH:MM'
            
            sheet.cell(row=row_num, column=8).value = 'Oui' if doc.get('validated') else 'Non'
            
            # Couleur selon statut
            status_cell = sheet.cell(row=row_num, column=5)
            if doc.get('validated'):
                status_cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
            else:
                status_cell.fill = PatternFill(start_color='FFECB3', end_color='FFECB3', fill_type='solid')
            
            # Bordures
            for col in range(1, 9):
                sheet.cell(row=row_num, column=col).border = self.border
        
        # Auto-ajuster
        for col in range(1, 9):
            sheet.column_dimensions[get_column_letter(col)].width = 15
    
    def add_summary_sheet(self, stats: Dict[str, Any]):
        """Ajoute une feuille de résumé avec statistiques"""
        sheet = self.create_sheet("Résumé", activate=False)
        
        # Titre
        sheet['A1'] = self.title
        sheet['A1'].font = Font(size=18, bold=True, color='673AB7')
        sheet.merge_cells('A1:D1')
        
        # Date
        sheet['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        sheet['A2'].font = Font(size=10, italic=True)
        
        # Statistiques
        row = 4
        for key, value in stats.items():
            sheet.cell(row=row, column=1).value = key
            sheet.cell(row=row, column=1).font = Font(bold=True)
            sheet.cell(row=row, column=2).value = value
            row += 1
        
        # Auto-ajuster
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 20
    
    def add_chart(
        self,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        title: str,
        position: str = 'E2'
    ):
        """
        Ajoute un graphique à une feuille
        
        Args:
            sheet_name: Nom de la feuille
            chart_type: Type ('bar', 'pie', 'line')
            data_range: Plage de données (ex: 'A1:B10')
            title: Titre du graphique
            position: Position (ex: 'E2')
        """
        sheet = self.wb[sheet_name]
        
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'pie':
            chart = PieChart()
        elif chart_type == 'line':
            chart = LineChart()
        else:
            return
        
        chart.title = title
        chart.style = 10
        
        # Ajouter les données (simplifié)
        # Note: Nécessite configuration plus détaillée en production
        sheet.add_chart(chart, position)
    
    def build(self) -> bytes:
        """
        Construit le fichier Excel et retourne les bytes
        
        Returns:
            bytes: Contenu du fichier Excel
        """
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def save(self, filename: str):
        """Sauvegarde le fichier Excel"""
        self.wb.save(filename)
